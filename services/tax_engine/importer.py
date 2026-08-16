from __future__ import annotations

import hashlib
import re
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .models import TaxRevenueEvent

PERIOD = re.compile(r"^[0-9]{4}-(0[1-9]|1[0-2])$")
REQUIRED = {"source_period","gross_realised","returns","partner_loyalty_payments","partner_loyalty_reversals"}
REPORT_PERIOD = re.compile(r"за период с (\d{2})\.(\d{2})\.(\d{4}) по (\d{2})\.(\d{2})\.(\d{4})", re.IGNORECASE)


class ImportContractError(ValueError):
    pass


def source_checksum(path: str | Path) -> str:
    digest = hashlib.sha256()
    with Path(path).open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _amount(value: Any, field: str) -> Decimal:
    try:
        result = Decimal(str(value).replace(",", "."))
    except (InvalidOperation, TypeError, ValueError) as error:
        raise ImportContractError(f"{field} must be numeric") from error
    if result < 0:
        raise ImportContractError(f"{field} must be non-negative")
    return result.quantize(Decimal("0.01"))


def _source_path(source_path: str | Path) -> Path:
    path = Path(source_path)
    if path.suffix.lower() not in {".xlsx", ".xls"} or not path.is_file():
        raise ImportContractError("an existing official Excel realization report is required")
    return path


def _find_exact(values: list[Any], label: str, start: int, end: int) -> int:
    matches = [index for index in range(start, end + 1) if values[index - 1] == label]
    if len(matches) != 1:
        raise ImportContractError(f"expected one {label!r} column in report group")
    return matches[0]


def extract_realization_workbook(source_path: str | Path) -> dict[str, Any]:
    """Read confirmed Ozon realization workbook structure without changing the source file."""
    path = _source_path(source_path)
    try:
        workbook = load_workbook(path, read_only=False, data_only=True)
    except Exception as error:
        raise ImportContractError("official realization workbook is not readable") from error
    try:
        if len(workbook.sheetnames) != 1:
            raise ImportContractError("realization workbook must contain exactly one report sheet")
        sheet = workbook[workbook.sheetnames[0]]
        header_rows = [row for row in range(1, sheet.max_row + 1) if sheet.cell(row, 1).value == "№ п/п"]
        if len(header_rows) != 1:
            raise ImportContractError("report header row was not found unambiguously")
        header_row = header_rows[0]
        top = [sheet.cell(header_row, column).value for column in range(1, sheet.max_column + 1)]
        sub = [sheet.cell(header_row + 1, column).value for column in range(1, sheet.max_column + 1)]
        realised_start = _find_exact(top, "Реализовано", 1, sheet.max_column)
        returned_start = _find_exact(top, "Возвращено клиентом", 1, sheet.max_column)
        later_groups = [index for index, value in enumerate(top, 1) if index > returned_start and value is not None]
        realised_end = returned_start - 1
        returned_end = min(later_groups) - 1 if later_groups else sheet.max_column
        columns = {
            "gross_realised": _find_exact(sub, "Реализовано на сумму, руб.", realised_start, realised_end),
            "partner_loyalty_payments": _find_exact(sub, "Выплаты по механикам лояльности партнёров, руб.", realised_start, realised_end),
            "realised_quantity": _find_exact(sub, "Кол-во", realised_start, realised_end),
            "returns": _find_exact(sub, "Возвращено на сумму, руб.", returned_start, returned_end),
            "partner_loyalty_reversals": _find_exact(sub, "Выплаты по механикам лояльности партнёров, руб.", returned_start, returned_end),
            "returned_quantity": _find_exact(sub, "Кол-во", returned_start, returned_end),
        }
        data_rows = [row for row in range(header_row + 3, sheet.max_row + 1)
                     if isinstance(sheet.cell(row, 1).value, (int, float))]
        if not data_rows:
            raise ImportContractError("realization workbook has no data rows")
        totals = {
            field: sum((_amount(sheet.cell(row, column).value or 0, field) for row in data_rows), Decimal("0"))
            for field, column in columns.items()
        }
        period_matches = []
        for row in range(1, header_row):
            for column in range(1, sheet.max_column + 1):
                value = sheet.cell(row, column).value
                if isinstance(value, str):
                    match = REPORT_PERIOD.search(value)
                    if match:
                        period_matches.append(match.groups())
        if len(period_matches) != 1:
            raise ImportContractError("report period was not found unambiguously")
        start_day, start_month, start_year, end_day, end_month, end_year = period_matches[0]
        if (start_year, start_month) != (end_year, end_month):
            raise ImportContractError("realization report spans more than one source period")
        source_period = f"{start_year}-{start_month}"
        report_kind = "ORDER_LEVEL" if "Отправление" in top else "MONTHLY"
        result = {
            "source_document": path.name,
            "source_checksum": source_checksum(path),
            "source_period": source_period,
            "report_kind": report_kind,
            "sheet_names": workbook.sheetnames,
            "header_row": header_row,
            "data_row_start": min(data_rows),
            "data_row_end": max(data_rows),
            "data_rows_count": len(data_rows),
            "actual_fields": {field: sub[column - 1] for field, column in columns.items()},
            "totals": totals,
        }
        if report_kind == "ORDER_LEVEL":
            shipment_start = _find_exact(top, "Отправление", 1, sheet.max_column)
            following = [index for index, value in enumerate(top, 1) if index > shipment_start and value is not None]
            shipment_end = min(following) - 1 if following else sheet.max_column
            number_column = _find_exact(sub, "Номер", shipment_start, shipment_end)
            date_column = _find_exact(sub, "Дата", shipment_start, shipment_end)
            dates = [sheet.cell(row, date_column).value for row in data_rows if sheet.cell(row, date_column).value is not None]
            result["order_fields"] = {"posting_number": sub[number_column - 1], "source_event_date": sub[date_column - 1]}
            result["source_event_date_min"] = min(dates) if dates else None
            result["source_event_date_max"] = max(dates) if dates else None
        return result
    finally:
        workbook.close()


def preview_monthly_workbook(source_path: str | Path) -> dict[str, Any]:
    extracted = extract_realization_workbook(source_path)
    if extracted["report_kind"] != "MONTHLY":
        raise ImportContractError("statutory preview requires the monthly realization workbook")
    totals = extracted["totals"]
    return preview_monthly_report({
        "source_period": extracted["source_period"],
        "gross_realised": totals["gross_realised"],
        "returns": totals["returns"],
        "partner_loyalty_payments": totals["partner_loyalty_payments"],
        "partner_loyalty_reversals": totals["partner_loyalty_reversals"],
    }, source_path)


def preview_monthly_report(document: dict[str, Any], source_path: str | Path) -> dict[str, Any]:
    """Normalize explicitly extracted official monthly totals; never infer Excel columns."""
    if not isinstance(document, dict) or not REQUIRED.issubset(document):
        raise ImportContractError("monthly realization contract is incomplete")
    period = str(document["source_period"])
    if not PERIOD.fullmatch(period):
        raise ImportContractError("source_period must be YYYY-MM")
    path = _source_path(source_path)
    checksum = source_checksum(path)
    reference = f"ozon-realization:{period}:{checksum}"
    components = [
        ("REALIZATION", _amount(document["gross_realised"], "gross_realised"), "CONFIRMED"),
        ("RETURN", -_amount(document["returns"], "returns"), "CONFIRMED"),
        ("PARTNER_LOYALTY_PAYMENT", _amount(document["partner_loyalty_payments"], "partner_loyalty_payments"), "PARTIAL"),
        ("PARTNER_LOYALTY_REVERSAL", -_amount(document["partner_loyalty_reversals"], "partner_loyalty_reversals"), "PARTIAL"),
    ]
    events = []
    for event_type, amount, semantics in components:
        key = f"{reference}|{event_type}|{amount}"
        events.append(TaxRevenueEvent(
            event_id=hashlib.sha256(key.encode()).hexdigest(), tax_year=int(period[:4]),
            source_period=period, event_type=event_type, amount=amount,
            source_document=path.name, source_reference=reference,
            tax_semantics_status=semantics, tax_date_status="PERIOD_ONLY",
            data_quality_status="valid" if semantics == "CONFIRMED" else "partial",
        ))
    return {"source_document":path.name,"source_reference":reference,"source_checksum":checksum,
            "source_period":period,"events":events,"accounting_income":sum((e.amount for e in events),Decimal("0")),
            "persist":False}
